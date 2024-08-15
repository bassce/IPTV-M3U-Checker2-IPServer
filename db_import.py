#!/usr/bin/env python
# -*- coding: utf-8 -*-
# not used,only test
import sqlite3
import openpyxl

def import_tvorders(dbconn=None, xlsfilename=r'playlists/sortlist.xlsx', bDeleteOld=True):
    try:
        conn = dbconn if (dbconn != None) else sqlite3.connect('database/db.sqlite3')
        try:
            conn.execute('''CREATE TABLE IF NOT EXISTS tvorders (
                                tvname NVARCHAR(30) NOT NULL PRIMARY KEY,
                                tvgroup NVARCHAR(30) NULL,
                                memo TEXT NULL,
                                tvorder INT NULL)''')
        except Exception as e:
            print(e)

        listinsheet = openpyxl.load_workbook(xlsfilename)
        datainlist = listinsheet.active
        if bDeleteOld:
            conn.execute("DELETE FROM tvorders")

        c = conn.cursor()
        data_truck = '''INSERT OR IGNORE INTO tvorders(tvname,tvgroup,memo,tvorder) VALUES (?,?,?,?)'''

        titles = set()  # 使用集合来确保唯一性
        for row in datainlist.iter_rows(min_row=2, max_col=4, max_row=datainlist.max_row):
            cargo = [cell.value for cell in row]
            if cargo[0] not in titles:
                titles.add(cargo[0])
                c.execute(data_truck, cargo)

        conn.commit()
        print("导入节目排序表成功！")
        
        if dbconn == None:
            conn.close()
        return datainlist.max_row - 1
    except Exception as e:
        print(e)
        return 0

if __name__ == '__main__':
    conn=sqlite3.connect('database\db.sqlite3') 
    if(0<import_tvorders(conn)):
        sql='select * from tvorders o order by o.tvorder'   #'select p.id,p.title,o.tvname,o.tvorder from playlists p left join tvorders o  on p.title=o.tvname order by p.tvgroup,o.tvorder '
        rows=conn.execute(sql)
        for row in rows:
            print(row)
    conn.close()